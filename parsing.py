import sqlite3
from tkinter import *
from tkinter.filedialog import askopenfilename
import os
from openpyxl import load_workbook

table_name = "parsing"  # Имя таблицы в базе данных
file_database = "data.db"  # Имя файла базы данных


def opening_a_file():
    """Открытие файла Excel"""
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    return filename


def parsing_document(min_row, max_row, column) -> None:
    """
    Осуществляет парсинг данных из файла Excel и вставляет их в базу данных SQLite.

    Аргументы:
    :param min_row: Строка, с которой начинается считывание данных.
    :param max_row: Строка, с которой заканчивается считывание данных.
    :param column: Столбец, с которого начинается считывание данных.
    """
    filename = opening_a_file()  # Открываем выбор файла Excel для чтения данных
    workbook = load_workbook(filename=filename)  # Загружаем выбранный файл Excel
    sheet = workbook.active

    os.remove(file_database)  # Удаляем файл базы данных

    conn = sqlite3.connect(file_database)  # Создаем соединение с базой данных
    cursor = conn.cursor()
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute(f"CREATE TABLE IF NOT EXISTS {table_name} (table_column_1)")
    # Считываем данные из колонки A и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=int(min_row), max_row=int(max_row), values_only=True):
        table_column_1 = str(row[int(column)])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute(f"SELECT * FROM {table_name} WHERE table_column_1 = ?", (table_column_1,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute(f"INSERT INTO {table_name} VALUES (?)", (table_column_1,))
    # Удаляем повторы по табельному номеру
    cursor.execute(f"DELETE FROM {table_name} WHERE rowid NOT IN (SELECT min(rowid) FROM {table_name} GROUP BY table_column_1)")
    # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def parsing_document_1(min_row, max_row, column, column_1) -> None:
    """
    Осуществляет парсинг данных из файла Excel и вставляет их в базу данных SQLite.

    Аргументы:
    :param min_row: Строка, с которой начинается считывание данных.
    :param max_row: Строка, с которой заканчивается считывание данных.
    :param column: Столбец, с которого начинается считывание данных.
    :param column_1: Столбец, с которого начинается считывание данных.
    """
    filename = opening_a_file()  # Открываем выбор файла Excel для чтения данных
    workbook = load_workbook(filename=filename)  # Загружаем выбранный файл Excel
    sheet = workbook.active

    os.remove(file_database)  # Удаляем файл базы данных

    conn = sqlite3.connect(file_database)  # Создаем соединение с базой данных
    cursor = conn.cursor()
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute(f"CREATE TABLE IF NOT EXISTS {table_name} (table_column_1, table_column_2)")
    # Считываем данные из колонки A и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=int(min_row), max_row=int(max_row), values_only=True):
        table_column_1 = str(row[int(column)])  # Преобразуем значение в строку
        table_column_2 = str(row[int(column_1)])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute(f"SELECT * FROM {table_name} WHERE table_column_1 = ? AND table_column_2 = ?",
                       (table_column_1, table_column_2))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute(f"INSERT INTO {table_name} VALUES (?, ?)", (table_column_1, table_column_2))
    # Удаляем повторы по табельному номеру
    cursor.execute(f"DELETE FROM {table_name} WHERE rowid NOT IN (SELECT min(rowid) FROM {table_name} GROUP BY table_column_1, table_column_2)")
    # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()
