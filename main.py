import sqlite3
from tkinter import *

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rich import print

from parsing import parsing_document, opening_a_file, table_name, parsing_document_1


def opening_the_database():
    """Открытие базы данных"""
    conn = sqlite3.connect('data.db')  # Создаем соединение с базой данных
    cursor = conn.cursor()
    return conn, cursor


def main():
    print("[bold red]Parsing всего! Давай Parsing все!\n",
          "[bold red][1] - Парсинг документа 1 колонка\n",
          "[bold red][2] - Парсинг документа 2 колонки\n",
          # "[bold red][2] - Сравниваем пенсионеров\n",
          # "[bold red][3] - Parsing профессии\n",
          "[bold red][3] - Сравниваем значение в колонке и в базе данных и если найдено совпадение, то записываем "
          "значение из базы данных (берем табельный номер из базы данных и сравниваем с табельным номером из файла, "
          "если совпадение найдено, то берем значение из базы данных равное табельному номеру например профессию и "
          "записываем в соответствующую колонку в файле)\n",
          # "[bold red][5] - Parsing ЗП Май 2023\n",
          # "[bold red][6] - Записываем ЗП Май 2023\n",
          # "[bold red][7] - Parsing ЗП Июнь 2023\n",
          # "[bold red][8] - Записываем ЗП Июнь 2023\n",
          # "[bold red][9] - Parsing ЗП Июнь 2023\n",
          # "[bold red][10] - Сравниваем ГО 2023\n",
          # "[bold red][11] - Parsing 10.23\n",
          # "[bold red][12] - Пометка цветом (подсчет колонок начинается с 1)\n"
          # "[bold red][13] - Парсинг данных 30.10.2023\n"
          # "[bold red][14] - Сравниваем имущество\n"
          # "[bold red][15] - Ищем дубликаты и помечаем определенным цветом\n"
          # "[bold red][16] - Ищем дубли по первому слову\n"
          # "[bold red][17] - Парсим данные в базу данных (имущество)"
          # "[bold red][18] - Сравниваем и записываем"
          )
    user_input = input("Сделай выбор: ")
    if user_input == "1":
        input_function(None, 'Введите номер строки, с которой начинается считывание данных:',
                       'Введите номер строки, с которой заканчивается считывание данных.:',
                       'Введите номер столбца, с которого начинается считывание данных.:')
    elif user_input == "2":
        input_function_1(None, 'Введите номер строки, с которой начинается считывание данных:',
                         'Введите номер строки, с которой заканчивается считывание данных.:',
                         'Введите номер столбца, с которого начинается считывание данных.:',
                         'Введите номер столбца, с которого начинается считывание данных.:')
    elif user_input == "3":
        compare_and_rewrite_professions()
    # elif user_input == "3":
    #     parsing_of_professions()
    # elif user_input == "4":
    #     compare_and_rewrite_professions()
    # elif user_input == "5":
    #     po_parsing_may_2023()
    # elif user_input == "6":
    #     compare_and_rewrite_professions_may_2023()
    # elif user_input == "7":
    #     po_parsing_jul_2023()
    # elif user_input == "8":
    #     compare_and_rewrite_professions_jul_2023()
    # elif user_input == "9":
    #     po_parsing_go_2023()
    # elif user_input == "10":
    #     comparing_the_data_go()  # Сравниваем данные с базы данных с файлом
    # elif user_input == "11":
    #     comparing_the_data_go_10_23()
    # elif user_input == "12":  # Сравниваем значения
    #     comparing_the_data_go_10_23_23(sheet_title='05.24', min_row=5, max_row=1072, column=3)
    # elif user_input == "13":
    #     property_parsing()
    # elif user_input == "14":
    #     comparing_property()  # Сравниваем имущество
    # elif user_input == "15":
    #     find_and_highlight_duplicates(filename='input_doc/Расчетная-ведомость за апрель - июнь 2024.xlsx',
    #                                   sheet_name='05.24')
    # elif user_input == "16":  # Поиск дубликатов по первому слову
    #     find_and_highlight_duplicates_by_first_word(
    #         filename='Шаблон ОДИ испр. (МУЭ тлг.5463) техотдел исправлено название.xlsx',
    #         sheet_name='T')
    # elif user_input == "17":
    #     analysis_of_the_completed_table(filename='Перечень ОНИ Минстрой (для Даши) Захаров.xlsx', sheet_name='шаблон')
    # elif user_input == "18":
    #     compare_and_write_down(filename='ОНИ 29.10.2023.xlsx')


def input_function(root, label1, label2, label3):
    """
    Графическое окно ввода
    :param root: Графическое окно
    :param label1: Поле ввода
    :param label2: Поле ввода
    :param label3: Поле ввода
    """
    # Создаем окно Tkinter
    root = Tk()
    root.geometry('450x180')  # Устанавливаем размер окна

    # Создаем два поля ввода
    entry1 = Entry(root)
    entry2 = Entry(root)
    entry3 = Entry(root)

    # Добавляем метки к полям ввода
    Label(root, text=label1).grid(row=0, column=0)
    Label(root, text=label2).grid(row=2, column=0)
    Label(root, text=label3).grid(row=4, column=0)

    # Размещаем поля ввода
    entry1.grid(row=1, column=0)
    entry2.grid(row=3, column=0)
    entry3.grid(row=5, column=0)

    # Увеличиваем ширину полей ввода
    entry1.config(width=70)  # Ширина поля ввода в символах
    entry2.config(width=70)
    entry3.config(width=70)

    # Создаем кнопку и добавляем обработчик событий
    Button(root, text="Готово", command=lambda: handle_done_button(entry1, entry2, entry3)).grid(row=7, column=0)

    # Запускаем цикл обработки событий
    root.mainloop()


def handle_done_button(entry1, entry2, entry3):
    print("Данные введены:", entry1.get(), entry2.get(), entry3.get())
    parsing_document(entry1.get(), entry2.get(), entry3.get())


def input_function_1(root, label1, label2, label3, label4):
    """
    Графическое окно ввода
    :param root: Графическое окно
    :param label1: Поле ввода
    :param label2: Поле ввода
    :param label3: Поле ввода
    """
    # Создаем окно Tkinter
    root = Tk()
    root.geometry('450x200')  # Устанавливаем размер окна

    # Создаем два поля ввода
    entry1 = Entry(root)
    entry2 = Entry(root)
    entry3 = Entry(root)
    entry4 = Entry(root)

    # Добавляем метки к полям ввода
    Label(root, text=label1).grid(row=0, column=0)
    Label(root, text=label2).grid(row=2, column=0)
    Label(root, text=label3).grid(row=4, column=0)
    Label(root, text=label4).grid(row=6, column=0)

    # Размещаем поля ввода
    entry1.grid(row=1, column=0)
    entry2.grid(row=3, column=0)
    entry3.grid(row=5, column=0)
    entry4.grid(row=7, column=0)

    # Увеличиваем ширину полей ввода
    entry1.config(width=70)  # Ширина поля ввода в символах
    entry2.config(width=70)
    entry3.config(width=70)
    entry4.config(width=70)

    # Создаем кнопку и добавляем обработчик событий
    Button(root, text="Готово", command=lambda: handle_done_button_1(entry1, entry2, entry3, entry4)).grid(row=8,
                                                                                                           column=0)

    # Запускаем цикл обработки событий
    root.mainloop()


def handle_done_button_1(entry1, entry2, entry3, entry4):
    print("Данные введены:", entry1.get(), entry2.get(), entry3.get(), entry4.get())
    parsing_document_1(entry1.get(), entry2.get(), entry3.get(), entry4.get())


def comparing_the_data_go_10_23():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_go_10_23 (service_number)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    min_row_input = input("Введите номер строки с которой будем parsing:")
    max_row_input = input("Введите номер строки до которой будем parsing:")
    st_row_input = input("Введите номер столбца будем parsing, счет начинается с 0:")
    for row in sheet.iter_rows(min_row=int(min_row_input), max_row=int(max_row_input), values_only=True):
        service_number = str(row[int(st_row_input)])  # Преобразуем значение в строку
        # zp = str(row[1])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_go_10_23 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_go_10_23 VALUES (?)', (service_number,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def download_excel_file(filename, sheet_name):
    """Загружает файл Excel"""
    workbook = load_workbook(filename)
    sheet = workbook[sheet_name]
    return sheet, workbook


def find_and_highlight_duplicates(filename, sheet_name):
    sheet, workbook = download_excel_file(filename, sheet_name)
    # Создаем множество для хранения уникальных значений
    unique_values = set()
    duplicates = set()
    # Задаем стиль подсветки для дубликатов
    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    # Поиск дубликатов и подсветка
    for row in range(5, 971):  # Проходим по строкам с 2 по 1542
        cell_value = sheet.cell(row=row, column=2).value  # Столбец 2 (считая с 0)
        print(cell_value)
        if cell_value in unique_values:
            duplicates.add(cell_value)
        else:
            unique_values.add(cell_value)
    # Подсветка дубликатов
    for row in range(2, 1543):
        cell_value = sheet.cell(row=row, column=3).value
        if cell_value in duplicates:
            sheet.cell(row=row, column=3).fill = fill
    # Сохранение изменений в файле
    workbook.save(filename)
    workbook.close()


def po_parsing_jul_2023():
    """Изменение от 19.01.2024 Парсинг май 2023 счет начинается с 0"""
    conn, cursor = opening_the_database()
    filename = opening_a_file()  # Открываем выбор файла Excel для чтения данных
    workbook = load_workbook(filename=filename)  # Загружаем выбранный файл Excel
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute(f"CREATE TABLE IF NOT EXISTS {table_name} (service_number, zp, otpysk)")
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=1, max_row=908, values_only=True):
        service_number = str(row[0])  # Преобразуем значение в строку
        zp = str(row[2])  # Преобразуем значение в строку
        otpysk = str(row[3])  # Преобразуем значение в строку
        print(service_number, zp)
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute(f'SELECT * FROM {table_name} WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute(f'INSERT INTO {table_name} VALUES (?, ?, ?)', (
                service_number,
                zp,
                otpysk,
            ))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def comparing_the_data_go_10_23_23(sheet_title, min_row, max_row, column):
    """
    Сравниваем данные с базы данных с файлом

    :param sheet_title: название вкладки
    :param min_row: строка с которой начинается считывание данных
    :param max_row: строка на которой заканчивается парсинг
    :param column: Столбец, с которого начинается считывание данных
    """

    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    workbook = load_workbook(filename=filename)  # Загружаем выбранный файл Excel
    # Получаем нужный лист по названию
    try:
        sheet = workbook[sheet_title]
    except KeyError:
        print(f"Лист с названием '{sheet_title}' не найден в файле.")
        return
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    # Читаем данные из базы данных и создаем множество с табельными номерами
    cursor.execute(f"SELECT service_number FROM {table_name}")
    db_service_numbers = {row[0] for row in cursor.fetchall()}
    # Считываем данные из указанного листа и вставляем их в базу данных
    for row_num, row in enumerate(
            sheet.iter_rows(min_row=int(min_row), max_row=int(max_row), values_only=True),
            start=int(min_row)):
        service_number = str(row[int(column)])
        print(service_number)
        # Проверяем, есть ли табельный номер в базе данных
        if service_number in db_service_numbers:
            for cell in sheet[row_num]:
                cell.fill = red_fill
    # Сохраняем изменения в файле
    workbook.save(filename)
    # Закрываем соединение с базой данных
    conn.commit()
    conn.close()


def property_parsing():
    """Парсинг имущества"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    workbook = load_workbook(filename='РЕЕСТР ОС ИТОГ 28 05 23 04-09.xlsx')  # Загружаем выбранный файл Excel
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute(
        '''CREATE TABLE IF NOT EXISTS property_parsing (service_number, zp, otpysk)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=int(7), max_row=int(756), values_only=True):
        number = str(row[int(2)])  # Преобразуем значение в строку Инвентарный номер
        area = str(row[int(15)])  # Площадь
        number_of_floors = str(row[int(16)])  # Количество этажей
        underground_floors = str(row[int(17)])  # Подземные этажи
        purpose_of_the_structure = str(row[int(22)])  # Назначение сооружения
        main_features = str(row[int(23)])  # Основные характеристики
        area_s = str(row[int(24)])  # Площадь
        length = str(row[int(25)])  # Протяженность
        depth = str(row[int(26)])  # Глубина
        depth_of_occurrence = str(row[int(27)])  # Глубина залегания
        height = str(row[int(28)])  # Высота
        volume = str(row[int(29)])  # Объем
        number_of_floors_s = str(row[int(30)])  # Колличество этажей
        main_features_s = str(row[int(31)])  # В том числе подземные
        cursor.execute('SELECT * FROM property_parsing WHERE number = ?', (number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            # Дописываем "0" в начале номера, чтобы он всегда состоял из 9 знаков
            number = number.zfill(9)
            cursor.execute('INSERT INTO property_parsing VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                           (number, area, number_of_floors, underground_floors,
                            purpose_of_the_structure,
                            main_features,
                            area_s,
                            length,
                            depth,
                            depth_of_occurrence,
                            height,
                            volume,
                            number_of_floors_s,
                            main_features_s))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def comparing_property():
    """Сравниваем данные с базы данных с файлом счет начинается 0"""
    conn, cursor = opening_the_database()
    # Загружаем файл Excel для записи результатов
    result_workbook = load_workbook(filename='input_doc/Состав ФОТ Апрель - Июнь_1.xlsx')
    result_sheet = result_workbook.active
    cursor.execute('SELECT service_number, zp, otpysk FROM parsing')  # Получаем все данные из базы данных
    db_data = cursor.fetchall()  # Получаем все записи из базы данных
    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонки G, H и I
    for row in result_sheet.iter_rows(min_row=5, max_row=794):
        value_D = str(row[0].value)  # Значение в колонке D
        print(value_D)
        db_number_list = [db_row for db_row in db_data if db_row[0] == value_D]
        print(db_number_list)
        if db_number_list:
            service_number = db_number_list[0][0]
            row[0].value = service_number  # Год из базы данных в колонку 20
            zp = db_number_list[0][1]
            row[22].value = zp
            otpysk = db_number_list[0][2]
            row[23].value = otpysk
        else:
            print("Не найдено")
    # Сохраняем изменения в файле Excel для записи результатов
    result_workbook.save(filename='input_doc/Состав ФОТ Апрель - Июнь_1.xlsx')
    result_workbook.close()


def compare_and_write_down(filename):
    """Сравниваем и записываем"""
    conn, cursor = opening_the_database()
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    cursor.execute('SELECT number, a, b, c, d, i FROM property_parsing')  # Получаем все данные из базы данных
    db_data = cursor.fetchall()  # Получаем все записи из базы данных
    # current_row = 3  # Start from the 3rd row
    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонки G, H и I
    for row in sheet.iter_rows(min_row=3, max_row=282):
        value_D = str(row[19].value)  # Значение в колонке D
        print(value_D)
        db_number_list = [db_row for db_row in db_data if db_row[0] == value_D]
        print(db_number_list)
        if db_number_list:
            a = db_number_list[0][1]
            # sheet.cell(row=current_row, column=4, value=a)
            row[3].value = a  # Год из базы данных в колонку 20
            b = db_number_list[0][2]
            row[5].value = b
            c = db_number_list[0][3]
            row[7].value = c
            d = db_number_list[0][4]
            row[9].value = d
            i = db_number_list[0][5]
            row[27].value = i
            # current_row += 1  # Увеличиваем номер строки
    # Сохраняем изменения в файле Excel для записи результатов
    workbook.save(filename)
    workbook.close()


def analysis_of_the_completed_table(filename, sheet_name):
    """Парсинг движемого и не движемого имущества"""
    conn, cursor = opening_the_database()
    # Загрузка файла Excel
    workbook = load_workbook(filename)
    sheet = workbook[sheet_name]
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS property_parsing (number, a, b, c, d, i)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    # ?min_row_input = input("Введите номер строки с которой будем parsing:")
    # max_row_input = input("Введите номер строки до которой будем parsing:")
    for row in sheet.iter_rows(min_row=int(5), max_row=int(483), values_only=True):
        number = str(row[int(15)])  # Преобразуем значение в строку Инвентарный номер
        a = str(row[int(2)])  # Дата ввода в эксплуатацию
        b = str(row[int(3)])  # Преобразуем значение в строку площадь кв.м
        c = str(row[int(4)])
        d = str(row[int(5)])
        i = str(row[int(17)])
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM property_parsing WHERE number = ?', (number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            # Дописываем "0" в начале номера, чтобы он всегда состоял из 9 знаков
            # number = number.zfill(9)
            cursor.execute('INSERT INTO property_parsing VALUES (?, ?, ?, ?, ?, ?)', (number, a, b, c, d, i))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def find_and_highlight_duplicates_by_first_word(filename, sheet_name):
    # Загрузка файла Excel
    workbook = load_workbook(filename)
    sheet = workbook[sheet_name]

    # Создаем множество для хранения уникальных значений первых слов
    unique_first_words = set()
    duplicates = set()

    # Задаем стиль подсветки для дубликатов
    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Поиск дубликатов и подсветка
    for row in range(2, 1543):  # Проходим по строкам с 2 по 1542
        cell_value = sheet.cell(row=row, column=3).value  # Столбец 2 (считая с 0)
        first_word = cell_value.split()[0] if cell_value else ''  # Получаем первое слово

        if first_word in unique_first_words:
            duplicates.add(first_word)
        else:
            unique_first_words.add(first_word)

    # Подсветка дубликатов
    for row in range(2, 1543):
        cell_value = sheet.cell(row=row, column=3).value
        first_word = cell_value.split()[0] if cell_value else ''

        if first_word in duplicates:
            sheet.cell(row=row, column=3).fill = fill

    # Сохранение изменений в файле
    workbook.save(filename)
    workbook.close()


def comparing_the_data_go():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем файл Excel для записи результатов
    result_workbook = load_workbook(filename=filename)
    result_sheet = result_workbook.active
    # Получаем все данные из базы данных
    cursor.execute('SELECT service_number FROM po_parsing_go_2023')
    db_data = [str(row[0]) for row in cursor.fetchall()]  # Преобразуем данные из базы данных в список строк
    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонку G
    for row in result_sheet.iter_rows(min_row=4, max_row=54):
        value_D = str(row[4].value)  # Значение в колонке D
        if value_D in db_data:
            row[12].value = 'Служит'  # Записываем 'пенсионер' в колонку G
    # Сохраняем изменения в файле Excel для записи результатов
    result_workbook.save(filename)
    result_workbook.close()


def po_parsing_go_2023():
    """Парсинг ГО"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_go_2023 (service_number TEXT PRIMARY KEY, zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=1, max_row=126, values_only=True):
        service_number = str(row[5])  # Преобразуем значение в строку
        zp = str(row[6])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_go_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_go_2023 VALUES (?, ?)', (service_number, zp,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions_jul_2023():
    """Сравнение и перезапись значений профессии в файле Excel"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM po_parsing_jul_2023')
    db_data = cursor.fetchall()
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1250):
        value_D = str(row[4].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[6].value = profession  # Записываем значение профессии в колонку C
    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()
    # Закрываем соединение с базой данных
    conn.close()


def compare_and_rewrite_professions_may_2023():
    """Сравнение и перезапись значений профессии в файле Excel"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM parsing')
    db_data = cursor.fetchall()
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=794):
        value_D = str(row[0].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[4].value = profession  # Записываем значение профессии в колонку C
    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()
    # Закрываем соединение с базой данных
    conn.close()


def po_parsing_may_2023():
    """Парсинг май 2023"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_may_2023 (service_number TEXT PRIMARY KEY, zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=11, max_row=1085, values_only=True):
        service_number = str(row[1])  # Преобразуем значение в строку
        zp = str(row[35])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_may_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_may_2023 VALUES (?, ?)', (service_number, zp,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions():
    """Изменение от 24.01.2024 Сравнение и перезапись значений профессии в файле Excel счет начинается с 0"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Считываем значения из базы данных
    cursor.execute(f'SELECT * FROM {table_name}')
    db_data = cursor.fetchall()
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=982):
        value_D = str(row[1].value)  # Значение в колонке с которой сравниваются данные
        print(value_D)
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[2].value = profession  # Записываем данные если найдены сходства

    workbook.save(filename)  # Сохраняем изменения в файле Excel
    workbook.close()
    conn.close()  # Закрываем соединение с базой данных


def parsing_of_professions():
    """Парсинг профессий"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS parsing (table_column_1, table_column_2)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=4, max_row=982, values_only=True):
        table_column_1 = str(row[1])  # Преобразуем значение в строку
        print(table_column_1)
        table_column_2 = str(row[2])  # Преобразуем значение в строку
        print(table_column_2)
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM parsing WHERE table_column_1 = ?', (table_column_1,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO parsing VALUES (?, ?)', (table_column_1, table_column_2,))
    # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def comparing_the_data():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем файл Excel для записи результатов
    result_workbook = load_workbook(filename=filename)
    result_sheet = result_workbook.active
    # Получаем все данные из базы данных
    cursor.execute('SELECT service_number FROM pensioners_zasyadko')
    db_data = [str(row[0]) for row in cursor.fetchall()]  # Преобразуем данные из базы данных в список строк
    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонку G
    for row in result_sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[2].value)  # Значение в колонке D
        if value_D in db_data:
            row[6].value = 'пенсионер'  # Записываем 'пенсионер' в колонку G
    # Сохраняем изменения в файле Excel для записи результатов
    result_workbook.save(filename)
    result_workbook.close()


if __name__ == "__main__":
    main()
