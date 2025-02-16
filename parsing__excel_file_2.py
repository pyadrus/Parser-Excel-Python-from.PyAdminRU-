import sqlite3
from tkinter import Tk
from tkinter.filedialog import askopenfilename

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rich import print


def opening_the_database():
    """Открытие базы данных"""
    conn = sqlite3.connect('data.db')  # Создаем соединение с базой данных
    cursor = conn.cursor()
    return conn, cursor


def opening_a_file():
    """Открытие файла Excel"""
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    return filename


def main():
    print("[bold red]Parsing всего! Давай Parsing все!\n",
          "[bold red][1] - Parsing пенсионеров\n",
          "[bold red][2] - Сравниваем пенсионеров\n",
          "[bold red][3] - Parsing профессии\n",
          "[bold red][4] - Сравниваем профессии\n",
          "[bold red][5] - Parsing ЗП Май 2023\n",
          "[bold red][6] - Записываем ЗП Май 2023\n",
          "[bold red][7] - Parsing ЗП Июнь 2023\n",
          "[bold red][8] - Записываем ЗП Июнь 2023\n",
          "[bold red][9] - Parsing ЗП Июнь 2023\n",
          "[bold red][10] - Сравниваем ГО 2023\n",
          "[bold red][11] - Parsing 10.23\n",
          "[bold red][12] - Пометка\n")
    user_input = input("Сделай выбор: ")
    if user_input == "1":
        parsing_pensioners()
    elif user_input == "2":
        comparing_the_data()
    elif user_input == "3":
        parsing_of_professions()
    elif user_input == "4":
        compare_and_rewrite_professions()
    elif user_input == "5":
        po_parsing_may_2023()
    elif user_input == "6":
        compare_and_rewrite_professions_may_2023()
    elif user_input == "7":
        po_parsing_jul_2023()
    elif user_input == "8":
        compare_and_rewrite_professions_jul_2023()
    elif user_input == "9":
        po_parsing_go_2023()
    elif user_input == "10":
        comparing_the_data_go()

    elif user_input == "11":
        comparing_the_data_go_10_23()
    elif user_input == "12":
        comparing_the_data_go_10_23_23()


def comparing_the_data_go_10_23_23():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Получаем название вкладки (листа) из ввода пользователя
    sheet_title = input("Введите название вкладки (листа) в файле Excel:")
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)

    # Получаем нужный лист по названию
    try:
        sheet = workbook[sheet_title]
    except KeyError:
        print(f"Лист с названием '{sheet_title}' не найден в файле.")
        return

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Читаем данные из базы данных и создаем множество с табельными номерами
    cursor.execute('SELECT service_number FROM po_parsing_go_10_23')
    db_service_numbers = {row[0] for row in cursor.fetchall()}

    # Считываем данные из указанного листа и вставляем их в базу данных
    min_row_input = input("Введите номер строки с которой будем parsing:")
    max_row_input = input("Введите номер строки до которой будем parsing:")
    st_col_input = input("Введите номер столбца будем parsing, счет начинается с 0:")

    for row_num, row in enumerate(
            sheet.iter_rows(min_row=int(min_row_input), max_row=int(max_row_input), values_only=True),
            start=int(min_row_input)):
        service_number = str(row[int(st_col_input)])

        # Проверяем, есть ли табельный номер в базе данных
        if service_number in db_service_numbers:
            for cell in sheet[row_num]:
                cell.fill = red_fill

    # Сохраняем изменения в файле
    workbook.save(filename)

    # Закрываем соединение с базой данных
    conn.commit()
    conn.close()


def comparing_the_data_go_10_23():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_go_10_23 (service_number)''')

    # Считываем данные из колонок A и H и вставляем их в базу данных
    min_row_input = input("Введите номер строки с которой будем parsing:")
    max_row_input = input("Введите номер строки до которой будем parsing:")
    st_row_input = input("Введите номер столбца будем parsing, счет начинается с 0:")

    for row in sheet.iter_rows(min_row=int(min_row_input), max_row=int(max_row_input), values_only=True):
        service_number = str(row[int(st_row_input)])  # Преобразуем значение в строку
        # zp = str(row[1])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_go_10_23 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_go_10_23 VALUES (?)', (service_number,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def comparing_the_data_go():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем файл Excel для записи результатов
    result_workbook = load_workbook(filename=filename)
    result_sheet = result_workbook.active
    # Получаем все данные из базы данных
    cursor.execute('SELECT service_number FROM po_parsing_go_2023')
    db_data = [str(row[0]) for row in cursor.fetchall()]  # Преобразуем данные из базы данных в список строк
    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонку G
    for row in result_sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[4].value)  # Значение в колонке D
        if value_D in db_data:
            row[12].value = 'Служит'  # Записываем 'пенсионер' в колонку G
    # Сохраняем изменения в файле Excel для записи результатов
    result_workbook.save(filename)
    result_workbook.close()


def po_parsing_go_2023():
    """Парсинг ГО"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_go_2023 (
                            service_number TEXT PRIMARY KEY,
                            zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=1, max_row=126, values_only=True):
        service_number = str(row[5])  # Преобразуем значение в строку
        zp = str(row[6])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_go_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_go_2023 VALUES (?, ?)', (service_number, zp,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions_jul_2023():
    """Сравнение и перезапись значений профессии в файле Excel"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM po_parsing_jul_2023')
    db_data = cursor.fetchall()
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1250):
        value_D = str(row[4].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[6].value = profession  # Записываем значение профессии в колонку C
    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()
    # Закрываем соединение с базой данных
    conn.close()


def po_parsing_jul_2023():
    """Парсинг май 2023"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_jul_2023 (
                            service_number TEXT PRIMARY KEY,
                            zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=12, max_row=1095, values_only=True):
        service_number = str(row[1])  # Преобразуем значение в строку
        zp = str(row[34])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_jul_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_jul_2023 VALUES (?, ?)', (service_number, zp,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions_may_2023():
    """Сравнение и перезапись значений профессии в файле Excel"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM po_parsing_may_2023')
    db_data = cursor.fetchall()
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[3].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[4].value = profession  # Записываем значение профессии в колонку C
    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()
    # Закрываем соединение с базой данных
    conn.close()


def po_parsing_may_2023():
    """Парсинг май 2023"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_may_2023 (
                            service_number TEXT PRIMARY KEY,
                            zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=11, max_row=1085, values_only=True):
        service_number = str(row[1])  # Преобразуем значение в строку
        zp = str(row[35])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_may_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_may_2023 VALUES (?, ?)', (service_number, zp,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions():
    """Сравнение и перезапись значений профессии в файле Excel"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM all_professions')
    db_data = cursor.fetchall()
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[3].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[2].value = profession  # Записываем значение профессии в колонку C
    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()
    # Закрываем соединение с базой данных
    conn.close()


def parsing_of_professions():
    """Парсинг профессий"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS all_professions (
                        service_number TEXT PRIMARY KEY,
                        professions TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=4, max_row=1249, values_only=True):
        service_number = str(row[0])  # Преобразуем значение в строку
        professions = str(row[7])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM all_professions WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO all_professions VALUES (?, ?)', (service_number, professions,))
    # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def comparing_the_data():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем файл Excel для записи результатов
    result_workbook = load_workbook(filename=filename)
    result_sheet = result_workbook.active
    # Получаем все данные из базы данных
    cursor.execute('SELECT service_number FROM pensioners_zasyadko')
    db_data = [str(row[0]) for row in cursor.fetchall()]  # Преобразуем данные из базы данных в список строк
    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонку G
    for row in result_sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[3].value)  # Значение в колонке D
        if value_D in db_data:
            row[6].value = 'пенсионер'  # Записываем 'пенсионер' в колонку G
    # Сохраняем изменения в файле Excel для записи результатов
    result_workbook.save(filename)
    result_workbook.close()


def parsing_pensioners():
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Остальной код остается без изменений...
    # Создаем соединение с базой данных
    conn = sqlite3.connect('data.db')
    cursor = conn.cursor()
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS pensioners_zasyadko (
                        service_number TEXT PRIMARY KEY)''')
    # Считываем данные из колонки A и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=4, max_row=649, values_only=True):
        service_number = str(row[0])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM pensioners_zasyadko WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO pensioners_zasyadko VALUES (?)', (service_number,))
    # Удаляем повторы по табельному номеру
    cursor.execute(
        'DELETE FROM pensioners_zasyadko WHERE rowid NOT IN (SELECT min(rowid) FROM pensioners_zasyadko GROUP BY service_number)')
    # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


if __name__ == "__main__":
    main()
