import sqlite3

import openpyxl

# Создаем подключение к базе данных
conn = sqlite3.connect('mydatabase.db')
cursor = conn.cursor()

# Создаем таблицу, если она еще не существует
cursor.execute("CREATE TABLE IF NOT EXISTS mytable (tab_num INTEGER)")

# Сохраняем изменения в базе данных
conn.commit()

# Открываем файл Excel
workbook = openpyxl.load_workbook('Копия sv1120700.xlsx')
sheet = workbook.active

# Парсим необходимые данные
for row in sheet.iter_rows(min_row=1, values_only=True):
    tab_num = row[0]

    # Сохраняем данные в базе данных
    cursor.execute("INSERT INTO mytable (tab_num) VALUES (?)", (tab_num,))

# Сохраняем изменения в базе данных
conn.commit()

# Закрываем соединение с базой данных
conn.close()


def write_to_excel():
    # Открываем файл Excel
    workbook = openpyxl.load_workbook('табульки.xlsx')
    sheet = workbook.active

    # Создаем подключение к базе данных
    conn = sqlite3.connect('mydatabase.db')
    cursor = conn.cursor()

    # Читаем данные из базы данных и записываем их в файл Excel
    for row in sheet.iter_rows(min_row=2):
        cell_C = row[2].value  # Значение ячейки C в текущей строке
        cursor.execute("SELECT tab_num FROM mytable WHERE tab_num=?", (cell_C,))
        result = cursor.fetchone()  # Получаем первый найденный результат

        if result is not None:
            # Если найдено соответствие в базе данных, записываем данные в файл Excel
            row[1].value = 'ВТБ'  # Записываем значение в столбец B

    # Сохраняем изменения в файле Excel
    workbook.save('табульки.xlsx')

    # Закрываем соединение с базой данных
    conn.close()


write_to_excel()
