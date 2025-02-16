import os

from openpyxl import load_workbook


def creating_folders():
    """Создание папок"""
    filename = "Списочный_состав.xlsx"

    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    min_row_input = 6  # Начальная строка
    max_row_input = 1071  # Конечная строка
    for row in sheet.iter_rows(min_row=int(min_row_input), max_row=int(max_row_input), values_only=True):
        name_rab = str(row[int(6)])  # Преобразуем значение в строку
        serv_namb = str(row[int(5)])  # Преобразуем значение в строку
        folder_name = f"{name_rab} {serv_namb}"

        os.makedirs(f"Список_людей/{folder_name}")
        print(folder_name)


if __name__ == "__main__":
    creating_folders()
