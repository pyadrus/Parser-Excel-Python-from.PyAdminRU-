import sqlite3

import openpyxl


def create_database_and_table():
    """
    Создает базу данных и таблицу для хранения информации о работниках и их часах работы.
    """
    conn = sqlite3.connect('mydatabase.db')
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mytable (
            tab_number INTEGER PRIMARY KEY,
            name TEXT,
            profession TEXT,
            work_time INTEGER
        )
    """)
    conn.commit()
    conn.close()


def read_data_from_excel_and_save_to_db(file_path):
    """
    Читает данные из Excel-файла и сохраняет их в базу данных SQLite.

    Параметры:
    - file_path (str): Путь к файлу Excel с данными о работниках.
    """
    try:
        # Подключаемся к базе данных
        conn = sqlite3.connect('mydatabase.db')
        cur = conn.cursor()

        # Загружаем Excel-файл
        workbook = openpyxl.load_workbook(filename=file_path)
        worksheet = workbook.active

        # Список для хранения данных
        data = []

        # Проходим по строкам файла Excel, начиная с 5 строки
        for row in worksheet.iter_rows(min_row=5, values_only=True):
            # Извлекаем необходимые данные из строк
            tab_number, name, profession, work_time = row[1], row[2], row[3], round(row[6] or 0)
            if tab_number:  # Проверяем, что табельный номер не пустой
                data.append((tab_number, name, profession, work_time))

        # Сохраняем данные в базу данных
        cur.executemany("INSERT INTO mytable (tab_number, name, profession, work_time) VALUES (?, ?, ?, ?)", data)
        conn.commit()

        print("Данные успешно загружены в базу данных.")
    except FileNotFoundError:
        print(f"Ошибка: Файл '{file_path}' не найден.")
    except Exception as e:
        print(f"Произошла ошибка при загрузке данных: {e}")
    finally:
        conn.close()


def update_excel_with_work_time(file_path):
    """
    Обновляет файл Excel с информацией о часах работы сотрудников из базы данных.

    Параметры:
    - file_path (str): Путь к файлу Excel, который нужно обновить.
    """
    try:
        # Подключаемся к базе данных
        conn = sqlite3.connect('mydatabase.db')
        cursor = conn.cursor()

        # Получаем все записи из таблицы базы данных
        cursor.execute("SELECT tab_number, work_time FROM mytable")
        rows = cursor.fetchall()

        # Загружаем файл Excel
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['Май 2023']

        # Обновляем значения времени работы в файле Excel
        for db_row in rows:
            tab_number, work_time = db_row
            for i in range(11, sheet.max_row + 1):
                if sheet.cell(row=i, column=2).value == tab_number:
                    sheet.cell(row=i, column=7).value = work_time
                    print(f"Обновлено значение для табельного номера {tab_number}: {work_time} часов.")
                    break

        # Сохраняем изменения в файл Excel
        workbook.save(file_path)
        print("Файл Excel успешно обновлен.")

    except FileNotFoundError:
        print(f"Ошибка: Файл '{file_path}' не найден.")
    except Exception as e:
        print(f"Произошла ошибка при обновлении файла Excel: {e}")
    finally:
        conn.close()


def main():
    """Основная функция для выполнения программы."""
    # Создаем базу данных и таблицу
    create_database_and_table()

    # Читаем данные из Excel и сохраняем в базу данных
    read_data_from_excel_and_save_to_db('зп февраль 2023.xlsx')

    # Обновляем файл Excel с данными о часах работы
    update_excel_with_work_time('табульки.xlsx')


if __name__ == "__main__":
    main()
