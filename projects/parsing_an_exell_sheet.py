# -*- coding: utf-8 -*-
import openpyxl as op

filename = "Копия promsvsh20220930502065449.xlsx"

wb = op.load_workbook(filename, data_only=True)
sheet = wb.active
max_col = sheet.max_column
max_ro = sheet.max_row

for i in range(3, max_ro + 1):
    namm = sheet.cell(row=1, column=2).value
    namms = namm.replace(' ', '')
    ruk = sheet.cell(row=i, column=2).value
    prof = sheet.cell(row=i, column=1).value

    if not ruk:
        continue
    if prof == "ИТОГО раб. пов.:":
        continue
    if prof == "ИТОГО ИТР пов.:":
        continue
    if prof == "ИТОГО по АП им. ш. А.Ф.Засядько:":
        continue

    print(f"{namms} - {prof}: {ruk}")

for i in range(3, max_ro + 1):
    namm = sheet.cell(row=1, column=3).value
    namms = namm.replace(' ', '')
    ruk = sheet.cell(row=i, column=3).value
    prof = sheet.cell(row=i, column=1).value

    if not ruk:
        continue
    if prof == "ИТОГО раб. пов.:":
        continue
    if prof == "ИТОГО ИТР пов.:":
        continue
    if prof == "ИТОГО по АП им. ш. А.Ф.Засядько:":
        continue
    print(f"{namms} - {prof}: {ruk}")
