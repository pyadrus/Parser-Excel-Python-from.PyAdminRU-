from gigachat import GigaChat
from openpyxl import load_workbook
from rich import print


def main():
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename='Копия NAC554850157.xlsx')
    sheet = workbook.active

    current_row = 6  # Start from the 3rd row

    for row in sheet.iter_rows(min_row=6, max_row=234, values_only=True):
        number = str(row[1])  # Считываем значение в колонке
        print(number)

        # Используйте токен, полученный в личном кабинете из поля Авторизационные данные
        with GigaChat(
                credentials='ZGNjMGU3YzYtNjZhOC00MjYwLWI1ODctYTUxYzNjZWNmYmZkOjJkNGIxMWRhLTY5ZjUtNDE5YS1hMmExLWNkNjMzODJkNTRhMw==',
                # Replace with your actual credentials
                verify_ssl_certs=False) as giga:
            response = giga.chat(
                f"Определи по имени, фамилии, отчеству это мужчина или женщина. Ответ должен быть одним словом: {number}")
            result = response.choices[0].message.content
            print(f'{number}: {result}')

            # Set the value of the cell in the 26th column, current_row
            sheet.cell(row=current_row, column=15, value=result)

            current_row += 1  # Move to the next row

    # Сохраняем изменения в файле
    workbook.save(filename='Копия NAC554850157.xlsx')


if __name__ == '__main__':
    main()
