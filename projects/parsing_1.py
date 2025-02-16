import sqlite3
from tkinter import Tk
from tkinter.filedialog import askopenfilename

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rich import print

from database.database import opening_the_database


def opening_a_file():
    """Открытие файла Excel"""
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    return filename


def property_parsing():
    """Парсинг имущества"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    workbook = load_workbook(filename='РЕЕСТР ОС ИТОГ 28 05 23 04-09.xlsx')  # Загружаем выбранный файл Excel
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute(
        '''CREATE TABLE IF NOT EXISTS property_parsing (number, area, number_of_floors, underground_floors,
                            purpose_of_the_structure,
                            main_features,
                            area_s,
                            length,
                            depth,
                            depth_of_occurrence,
                            height,
                            volume,
                            number_of_floors_s,
                            main_features_s)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=int(7), max_row=int(756), values_only=True):
        number = str(row[int(2)])  # Преобразуем значение в строку Инвентарный номер
        area = str(row[int(15)])  # Площадь
        number_of_floors = str(row[int(16)])  # Количество этажей
        underground_floors = str(row[int(17)])  # Подземные этажи
        purpose_of_the_structure = str(row[int(22)])  # Назначение сооружения
        main_features = str(row[int(23)])  # Основные характеристики
        area_s = str(row[int(24)])  # Площадь
        length = str(row[int(25)])  # Протяженность
        depth = str(row[int(26)])  # Глубина
        depth_of_occurrence = str(row[int(27)])  # Глубина залегания
        height = str(row[int(28)])  # Высота
        volume = str(row[int(29)])  # Объем
        number_of_floors_s = str(row[int(30)])  # Колличество этажей
        main_features_s = str(row[int(31)])  # В том числе подземные
        cursor.execute('SELECT * FROM property_parsing WHERE number = ?', (number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            # Дописываем "0" в начале номера, чтобы он всегда состоял из 9 знаков
            number = number.zfill(9)
            cursor.execute('INSERT INTO property_parsing VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                           (number, area, number_of_floors, underground_floors,
                            purpose_of_the_structure,
                            main_features,
                            area_s,
                            length,
                            depth,
                            depth_of_occurrence,
                            height,
                            volume,
                            number_of_floors_s,
                            main_features_s))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def comparing_property():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Загружаем файл Excel для записи результатов
    result_workbook = load_workbook(filename='ОНИ 30.10.2023.xlsx')
    result_sheet = result_workbook.active
    cursor.execute(
        'SELECT number, area, number_of_floors, underground_floors, purpose_of_the_structure, main_features, area_s, length, depth, depth_of_occurrence, height, volume, number_of_floors_s, main_features_s FROM property_parsing')  # Получаем все данные из базы данных
    db_data = cursor.fetchall()  # Получаем все записи из базы данных
    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонки G, H и I
    for row in result_sheet.iter_rows(min_row=3, max_row=276):
        value_D = str(row[15].value)  # Значение в колонке D
        print(value_D)
        db_number_list = [db_row for db_row in db_data if db_row[0] == value_D]
        print(db_number_list)
        if db_number_list:
            area = db_number_list[0][1]
            row[30].value = area  # Год из базы данных в колонку 20
            number_of_floors = db_number_list[0][2]
            row[31].value = number_of_floors
            underground_floors = db_number_list[0][3]
            row[32].value = underground_floors
            purpose_of_the_structure = db_number_list[0][4]
            row[33].value = purpose_of_the_structure  # Год из базы данных в колонку 20
            main_features = db_number_list[0][5]
            row[34].value = main_features
            area_s = db_number_list[0][6]
            row[35].value = area_s
            length = db_number_list[0][7]
            row[36].value = length  # Год из базы данных в колонку 20
            depth = db_number_list[0][8]
            row[37].value = depth
            depth_of_occurrence = db_number_list[0][9]
            row[38].value = depth_of_occurrence
            height = db_number_list[0][10]
            row[39].value = height  # Год из базы данных в колонку 20
            volume = db_number_list[0][11]
            row[40].value = volume
            number_of_floors_s = db_number_list[0][12]
            row[41].value = number_of_floors_s
            main_features_s = db_number_list[0][13]
            row[42].value = main_features_s
        else:
            print("Не найдено")
    # Сохраняем изменения в файле Excel для записи результатов
    result_workbook.save(filename='ОНИ 30.10.2023.xlsx')
    result_workbook.close()


def main():
    print("[bold red]Parsing всего! Давай Parsing все!\n",
          "[bold red][1] - Parsing пенсионеров\n",
          "[bold red][2] - Сравниваем пенсионеров\n",
          "[bold red][3] - Parsing профессии\n",
          "[bold red][4] - Сравниваем профессии\n",
          "[bold red][5] - Parsing ЗП Май 2023\n",
          "[bold red][6] - Записываем ЗП Май 2023\n",
          "[bold red][7] - Parsing ЗП Июнь 2023\n",
          "[bold red][8] - Записываем ЗП Июнь 2023\n",
          "[bold red][9] - Parsing ЗП Июнь 2023\n",
          "[bold red][10] - Сравниваем ГО 2023\n",
          "[bold red][11] - Parsing 10.23\n",
          "[bold red][12] - Пометка\n"
          "[bold red][13] - Парсинг данных 30.10.2023\n"
          "[bold red][14] - Сравниваем имущество\n"
          "[bold red][15] - Ищем дубликаты\n"
          "[bold red][16] - Ищем дубли по первому слову\n"
          "[bold red][17] - Парсим данные в базу данных (имущество)"
          "[18] - Сравниваем и записываем")
    user_input = input("Сделай выбор: ")
    if user_input == "1":
        parsing_pensioners()
    elif user_input == "2":
        comparing_the_data()
    elif user_input == "3":
        parsing_of_professions()
    elif user_input == "4":
        compare_and_rewrite_professions()
    elif user_input == "5":
        po_parsing_may_2023()
    elif user_input == "6":
        compare_and_rewrite_professions_may_2023()
    elif user_input == "7":
        po_parsing_jul_2023()
    elif user_input == "8":
        compare_and_rewrite_professions_jul_2023()
    elif user_input == "9":
        po_parsing_go_2023()
    elif user_input == "10":
        comparing_the_data_go()
    elif user_input == "11":
        comparing_the_data_go_10_23()
    elif user_input == "12":
        comparing_the_data_go_10_23_23()
    elif user_input == "13":
        property_parsing()
    elif user_input == "14":
        comparing_property()  # Сравниваем имущество
    elif user_input == "15":
        find_and_highlight_duplicates(filename='Шаблон ОДИ испр. (МУЭ тлг.5463) техотдел исправлено название.xlsx',
                                      sheet_name='T')
    elif user_input == "16":
        find_and_highlight_duplicates_by_first_word(
            filename='Шаблон ОДИ испр. (МУЭ тлг.5463) техотдел исправлено название.xlsx',
            sheet_name='T')
    elif user_input == "17":
        analysis_of_the_completed_table(filename='Перечень ОНИ Минстрой (для Даши) Захаров.xlsx', sheet_name='шаблон')
    elif user_input == "18":
        compare_and_write_down(filename='ОНИ 29.10.2023.xlsx')


def compare_and_write_down(filename):
    """Сравниваем и записываем"""
    conn, cursor = opening_the_database()
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    cursor.execute('SELECT number, a, b, c, d, i FROM property_parsing')  # Получаем все данные из базы данных
    db_data = cursor.fetchall()  # Получаем все записи из базы данных
    # current_row = 3  # Start from the 3rd row
    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонки G, H и I
    for row in sheet.iter_rows(min_row=3, max_row=282):
        value_D = str(row[19].value)  # Значение в колонке D
        print(value_D)
        db_number_list = [db_row for db_row in db_data if db_row[0] == value_D]
        print(db_number_list)
        if db_number_list:
            a = db_number_list[0][1]
            # sheet.cell(row=current_row, column=4, value=a)
            row[3].value = a  # Год из базы данных в колонку 20
            b = db_number_list[0][2]
            row[5].value = b
            c = db_number_list[0][3]
            row[7].value = c
            d = db_number_list[0][4]
            row[9].value = d
            i = db_number_list[0][5]
            row[27].value = i
            # current_row += 1  # Увеличиваем номер строки
    # Сохраняем изменения в файле Excel для записи результатов
    workbook.save(filename)
    workbook.close()


def analysis_of_the_completed_table(filename, sheet_name):
    """Парсинг движемого и не движемого имущества"""
    conn, cursor = opening_the_database()
    # Загрузка файла Excel
    workbook = load_workbook(filename)
    sheet = workbook[sheet_name]
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS property_parsing (number, a, b, c, d, i)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    # ?min_row_input = input("Введите номер строки с которой будем parsing:")
    # max_row_input = input("Введите номер строки до которой будем parsing:")
    for row in sheet.iter_rows(min_row=int(5), max_row=int(483), values_only=True):
        number = str(row[int(15)])  # Преобразуем значение в строку Инвентарный номер
        a = str(row[int(2)])  # Дата ввода в эксплуатацию
        b = str(row[int(3)])  # Преобразуем значение в строку площадь кв.м
        c = str(row[int(4)])
        d = str(row[int(5)])
        i = str(row[int(17)])
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM property_parsing WHERE number = ?', (number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            # Дописываем "0" в начале номера, чтобы он всегда состоял из 9 знаков
            # number = number.zfill(9)
            cursor.execute('INSERT INTO property_parsing VALUES (?, ?, ?, ?, ?, ?)', (number, a, b, c, d, i))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def find_and_highlight_duplicates_by_first_word(filename, sheet_name):
    # Загрузка файла Excel
    workbook = load_workbook(filename)
    sheet = workbook[sheet_name]

    # Создаем множество для хранения уникальных значений первых слов
    unique_first_words = set()
    duplicates = set()

    # Задаем стиль подсветки для дубликатов
    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Поиск дубликатов и подсветка
    for row in range(2, 1543):  # Проходим по строкам с 2 по 1542
        cell_value = sheet.cell(row=row, column=3).value  # Столбец 2 (считая с 0)
        first_word = cell_value.split()[0] if cell_value else ''  # Получаем первое слово

        if first_word in unique_first_words:
            duplicates.add(first_word)
        else:
            unique_first_words.add(first_word)

    # Подсветка дубликатов
    for row in range(2, 1543):
        cell_value = sheet.cell(row=row, column=3).value
        first_word = cell_value.split()[0] if cell_value else ''

        if first_word in duplicates:
            sheet.cell(row=row, column=3).fill = fill

    # Сохранение изменений в файле
    workbook.save(filename)
    workbook.close()


def find_and_highlight_duplicates(filename, sheet_name):
    # Загрузка файла Excel
    workbook = load_workbook(filename)
    sheet = workbook[sheet_name]

    # Создаем множество для хранения уникальных значений
    unique_values = set()
    duplicates = set()

    # Задаем стиль подсветки для дубликатов
    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Поиск дубликатов и подсветка
    for row in range(2, 1543):  # Проходим по строкам с 2 по 1542
        cell_value = sheet.cell(row=row, column=3).value  # Столбец 2 (считая с 0)
        if cell_value in unique_values:
            duplicates.add(cell_value)
        else:
            unique_values.add(cell_value)

    # Подсветка дубликатов
    for row in range(2, 1543):
        cell_value = sheet.cell(row=row, column=3).value
        if cell_value in duplicates:
            sheet.cell(row=row, column=3).fill = fill

    # Сохранение изменений в файле
    workbook.save(filename)
    workbook.close()


def comparing_the_data_go_10_23_23():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Получаем название вкладки (листа) из ввода пользователя
    sheet_title = input("Введите название вкладки (листа) в файле Excel:")
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    # Получаем нужный лист по названию
    try:
        sheet = workbook[sheet_title]
    except KeyError:
        print(f"Лист с названием '{sheet_title}' не найден в файле.")
        return
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    # Читаем данные из базы данных и создаем множество с табельными номерами
    cursor.execute('SELECT service_number FROM po_parsing_go_10_23')
    db_service_numbers = {row[0] for row in cursor.fetchall()}
    # Считываем данные из указанного листа и вставляем их в базу данных
    min_row_input = input("Введите номер строки с которой будем parsing:")
    max_row_input = input("Введите номер строки до которой будем parsing:")
    st_col_input = input("Введите номер столбца будем parsing, счет начинается с 0:")
    for row_num, row in enumerate(
            sheet.iter_rows(min_row=int(min_row_input), max_row=int(max_row_input), values_only=True),
            start=int(min_row_input)):
        service_number = str(row[int(st_col_input)])
        # Проверяем, есть ли табельный номер в базе данных
        if service_number in db_service_numbers:
            for cell in sheet[row_num]:
                cell.fill = red_fill
    # Сохраняем изменения в файле
    workbook.save(filename)
    # Закрываем соединение с базой данных
    conn.commit()
    conn.close()


def comparing_the_data_go_10_23():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_go_10_23 (service_number)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    min_row_input = input("Введите номер строки с которой будем parsing:")
    max_row_input = input("Введите номер строки до которой будем parsing:")
    st_row_input = input("Введите номер столбца будем parsing, счет начинается с 0:")
    for row in sheet.iter_rows(min_row=int(min_row_input), max_row=int(max_row_input), values_only=True):
        service_number = str(row[int(st_row_input)])  # Преобразуем значение в строку
        # zp = str(row[1])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_go_10_23 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_go_10_23 VALUES (?)', (service_number,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def comparing_the_data_go():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем файл Excel для записи результатов
    result_workbook = load_workbook(filename=filename)
    result_sheet = result_workbook.active
    # Получаем все данные из базы данных
    cursor.execute('SELECT service_number FROM po_parsing_go_2023')
    db_data = [str(row[0]) for row in cursor.fetchall()]  # Преобразуем данные из базы данных в список строк
    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонку G
    for row in result_sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[4].value)  # Значение в колонке D
        if value_D in db_data:
            row[12].value = 'Служит'  # Записываем 'пенсионер' в колонку G
    # Сохраняем изменения в файле Excel для записи результатов
    result_workbook.save(filename)
    result_workbook.close()


def po_parsing_go_2023():
    """Парсинг ГО"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_go_2023 (
                            service_number TEXT PRIMARY KEY,
                            zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=1, max_row=126, values_only=True):
        service_number = str(row[5])  # Преобразуем значение в строку
        zp = str(row[6])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_go_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_go_2023 VALUES (?, ?)', (service_number, zp,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions_jul_2023():
    """Сравнение и перезапись значений профессии в файле Excel"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM po_parsing_jul_2023')
    db_data = cursor.fetchall()
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1250):
        value_D = str(row[4].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[6].value = profession  # Записываем значение профессии в колонку C
    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()
    # Закрываем соединение с базой данных
    conn.close()


def po_parsing_jul_2023():
    """Парсинг май 2023"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_jul_2023 (
                            service_number TEXT PRIMARY KEY,
                            zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=12, max_row=1095, values_only=True):
        service_number = str(row[1])  # Преобразуем значение в строку
        zp = str(row[34])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_jul_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_jul_2023 VALUES (?, ?)', (service_number, zp,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions_may_2023():
    """Сравнение и перезапись значений профессии в файле Excel"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM po_parsing_may_2023')
    db_data = cursor.fetchall()
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[3].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[4].value = profession  # Записываем значение профессии в колонку C
    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()
    # Закрываем соединение с базой данных
    conn.close()


def po_parsing_may_2023():
    """Парсинг май 2023"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_may_2023 (
                            service_number TEXT PRIMARY KEY,
                            zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=11, max_row=1085, values_only=True):
        service_number = str(row[1])  # Преобразуем значение в строку
        zp = str(row[35])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_may_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_may_2023 VALUES (?, ?)', (service_number, zp,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions():
    """Сравнение и перезапись значений профессии в файле Excel"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM all_professions')
    db_data = cursor.fetchall()
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[3].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[2].value = profession  # Записываем значение профессии в колонку C
    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()
    # Закрываем соединение с базой данных
    conn.close()


def parsing_of_professions():
    """Парсинг профессий"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS all_professions (
                        service_number TEXT PRIMARY KEY,
                        professions TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=4, max_row=1249, values_only=True):
        service_number = str(row[0])  # Преобразуем значение в строку
        professions = str(row[7])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM all_professions WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO all_professions VALUES (?, ?)', (service_number, professions,))
    # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def comparing_the_data():
    """Сравниваем данные с базы данных с файлом"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем файл Excel для записи результатов
    result_workbook = load_workbook(filename=filename)
    result_sheet = result_workbook.active
    # Получаем все данные из базы данных
    cursor.execute('SELECT service_number FROM pensioners_zasyadko')
    db_data = [str(row[0]) for row in cursor.fetchall()]  # Преобразуем данные из базы данных в список строк
    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонку G
    for row in result_sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[3].value)  # Значение в колонке D
        if value_D in db_data:
            row[6].value = 'пенсионер'  # Записываем 'пенсионер' в колонку G
    # Сохраняем изменения в файле Excel для записи результатов
    result_workbook.save(filename)
    result_workbook.close()


def parsing_pensioners():
    # Открываем выбор файла Excel для чтения данных
    filename = opening_a_file()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Остальной код остается без изменений...
    # Создаем соединение с базой данных
    conn = sqlite3.connect('data.db')
    cursor = conn.cursor()
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS pensioners_zasyadko (
                        service_number TEXT PRIMARY KEY)''')
    # Считываем данные из колонки A и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=4, max_row=649, values_only=True):
        service_number = str(row[0])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM pensioners_zasyadko WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO pensioners_zasyadko VALUES (?)', (service_number,))
    # Удаляем повторы по табельному номеру
    cursor.execute(
        'DELETE FROM pensioners_zasyadko WHERE rowid NOT IN (SELECT min(rowid) FROM pensioners_zasyadko GROUP BY service_number)')
    # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


if __name__ == "__main__":
    main()
