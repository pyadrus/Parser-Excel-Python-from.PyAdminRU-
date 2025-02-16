# -*- coding: utf-8 -*-

from tkinter import Tk
from tkinter.filedialog import askopenfilename

from openpyxl import load_workbook

from database.database import opening_the_database


def main():
    print("[1] - Парсим пенсионеров")
    print("[2] - Сравниваем пенсионеров")
    print("[3] - Парсим профессии")
    print("[4] - Сравниваем профессии")
    print("[5] - Парсинг ЗП Май 2023")
    print("[6] - Записываем ЗП Май 2023")
    print("[7] - Парсинг ЗП Июнь 2023")
    print("[8] - Записываем ЗП Июнь 2023")
    print("[9] - Парсинг ЗП Июнь 2023")
    print("[10] - Сравниваем ГО 2023")
    user_input = input("Сделай выбор: ")
    if user_input == "1":
        parsing_pensioners()
    elif user_input == "2":
        comparing_the_data()
    elif user_input == "3":
        parsing_of_professions()
    elif user_input == "4":
        compare_and_rewrite_professions()
    elif user_input == "5":
        po_parsing_may_2023()
    elif user_input == "6":
        compare_and_rewrite_professions_may_2023()
    elif user_input == "7":
        po_parsing_jul_2023()
    elif user_input == "8":
        compare_and_rewrite_professions_jul_2023()
    elif user_input == "9":
        po_parsing_go_2023()
    elif user_input == "10":
        comparing_the_data_go()


def comparing_the_data_go():
    """Сравниваем данные с базы данных с файлом"""

    # Остальной код остается без изменений...
    # Создаем соединение с базой данных
    conn, cursor = opening_the_database()

    # Открываем выбор файла Excel для записи результатов
    root = Tk()
    root.withdraw()
    result_filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])

    # Проверяем, был ли выбран файл для записи результатов
    if not result_filename:
        print("Файл для записи результатов не выбран. Программа завершена.")
        exit()

    # Загружаем файл Excel для записи результатов
    result_workbook = load_workbook(filename=result_filename)
    result_sheet = result_workbook.active

    # Получаем все данные из базы данных
    cursor.execute('SELECT service_number FROM po_parsing_go_2023')
    db_data = [str(row[0]) for row in cursor.fetchall()]  # Преобразуем данные из базы данных в список строк

    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонку G
    for row in result_sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[4].value)  # Значение в колонке D
        if value_D in db_data:
            row[12].value = 'Служит'  # Записываем 'пенсионер' в колонку G

    # Сохраняем изменения в файле Excel для записи результатов
    result_workbook.save(result_filename)
    result_workbook.close()


def po_parsing_go_2023():
    """Парсинг ГО"""
    # Создаем соединение с базой данных
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    # Проверяем, был ли выбран файл
    if not filename:
        print("Файл не выбран. Программа завершена.")
        exit()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_go_2023 (
                            service_number TEXT PRIMARY KEY,
                            zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=1, max_row=126, values_only=True):
        service_number = str(row[5])  # Преобразуем значение в строку
        zp = str(row[6])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_go_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()

        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_go_2023 VALUES (?, ?)', (service_number, zp,))

        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions_jul_2023():
    """Сравнение и перезапись значений профессии в файле Excel"""

    # Создаем соединение с базой данных
    conn, cursor = opening_the_database()

    # Открываем выбор файла Excel для чтения данных
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])

    # Проверяем, был ли выбран файл
    if not filename:
        print("Файл не выбран. Программа завершена.")
        exit()

    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM po_parsing_jul_2023')
    db_data = cursor.fetchall()

    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1250):
        value_D = str(row[4].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[6].value = profession  # Записываем значение профессии в колонку C

    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()

    # Закрываем соединение с базой данных
    conn.close()


def po_parsing_jul_2023():
    """Парсинг май 2023"""
    # Создаем соединение с базой данных
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    # Проверяем, был ли выбран файл
    if not filename:
        print("Файл не выбран. Программа завершена.")
        exit()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_jul_2023 (
                            service_number TEXT PRIMARY KEY,
                            zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=12, max_row=1095, values_only=True):
        service_number = str(row[1])  # Преобразуем значение в строку
        zp = str(row[34])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_jul_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()

        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_jul_2023 VALUES (?, ?)', (service_number, zp,))

        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions_may_2023():
    """Сравнение и перезапись значений профессии в файле Excel"""

    # Создаем соединение с базой данных
    conn, cursor = opening_the_database()

    # Открываем выбор файла Excel для чтения данных
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])

    # Проверяем, был ли выбран файл
    if not filename:
        print("Файл не выбран. Программа завершена.")
        exit()

    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM po_parsing_may_2023')
    db_data = cursor.fetchall()

    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[3].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[4].value = profession  # Записываем значение профессии в колонку C

    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()

    # Закрываем соединение с базой данных
    conn.close()


def po_parsing_may_2023():
    """Парсинг май 2023"""
    # Создаем соединение с базой данных
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    # Проверяем, был ли выбран файл
    if not filename:
        print("Файл не выбран. Программа завершена.")
        exit()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS po_parsing_may_2023 (
                            service_number TEXT PRIMARY KEY,
                            zp TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=11, max_row=1085, values_only=True):
        service_number = str(row[1])  # Преобразуем значение в строку
        zp = str(row[35])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM po_parsing_may_2023 WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()

        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO po_parsing_may_2023 VALUES (?, ?)', (service_number, zp,))

        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def compare_and_rewrite_professions():
    """Сравнение и перезапись значений профессии в файле Excel"""

    # Создаем соединение с базой данных
    conn, cursor = opening_the_database()

    # Открываем выбор файла Excel для чтения данных
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])

    # Проверяем, был ли выбран файл
    if not filename:
        print("Файл не выбран. Программа завершена.")
        exit()

    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    # Считываем значения из базы данных
    cursor.execute('SELECT * FROM all_professions')
    db_data = cursor.fetchall()

    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[3].value)  # Значение в колонке D
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[2].value = profession  # Записываем значение профессии в колонку C

    # Сохраняем изменения в файле Excel
    workbook.save(filename)
    workbook.close()

    # Закрываем соединение с базой данных
    conn.close()


def parsing_of_professions():
    """Парсинг профессий"""
    # Создаем соединение с базой данных
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    # Проверяем, был ли выбран файл
    if not filename:
        print("Файл не выбран. Программа завершена.")
        exit()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS all_professions (
                        service_number TEXT PRIMARY KEY,
                        professions TEXT)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=4, max_row=1249, values_only=True):
        service_number = str(row[0])  # Преобразуем значение в строку
        professions = str(row[7])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM all_professions WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()

        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO all_professions VALUES (?, ?)', (service_number, professions,))

    # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def comparing_the_data():
    """Сравниваем данные с базы данных с файлом"""

    # Остальной код остается без изменений...
    # Создаем соединение с базой данных
    conn, cursor = opening_the_database()

    # Открываем выбор файла Excel для записи результатов
    root = Tk()
    root.withdraw()
    result_filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])

    # Проверяем, был ли выбран файл для записи результатов
    if not result_filename:
        print("Файл для записи результатов не выбран. Программа завершена.")
        exit()

    # Загружаем файл Excel для записи результатов
    result_workbook = load_workbook(filename=result_filename)
    result_sheet = result_workbook.active

    # Получаем все данные из базы данных
    cursor.execute('SELECT service_number FROM pensioners_zasyadko')
    db_data = [str(row[0]) for row in cursor.fetchall()]  # Преобразуем данные из базы данных в список строк

    # Сравниваем значения в колонке D с базой данных и записываем результаты в колонку G
    for row in result_sheet.iter_rows(min_row=5, max_row=1267):
        value_D = str(row[3].value)  # Значение в колонке D
        if value_D in db_data:
            row[6].value = 'пенсионер'  # Записываем 'пенсионер' в колонку G

    # Сохраняем изменения в файле Excel для записи результатов
    result_workbook.save(result_filename)
    result_workbook.close()


def parsing_pensioners():
    # Открываем диалоговое окно для выбора файла Excel
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])

    # Проверяем, был ли выбран файл
    if not filename:
        print("Файл не выбран. Программа завершена.")
        exit()

    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    # Остальной код остается без изменений...
    # Создаем соединение с базой данных
    conn, cursor = opening_the_database()

    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS pensioners_zasyadko (
                        service_number TEXT PRIMARY KEY)''')

    # Считываем данные из колонки A и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=4, max_row=649, values_only=True):
        service_number = str(row[0])  # Преобразуем значение в строку

        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM pensioners_zasyadko WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()

        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO pensioners_zasyadko VALUES (?)', (service_number,))

    # Удаляем повторы по табельному номеру
    cursor.execute(
        'DELETE FROM pensioners_zasyadko WHERE rowid NOT IN (SELECT min(rowid) FROM pensioners_zasyadko GROUP BY service_number)')

    # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


if __name__ == "__main__":
    main()
