# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook

from database.database import opening_the_database


def we_form_the_working_hours():
    """Формируем часы работы с файла переписываем в базу данных"""
    # Создаем базу данных и таблицу
    conn, cursor = opening_the_database()
    cursor.execute(
        "CREATE TABLE IF NOT EXISTS mytable (tab_number INTEGER, name TEXT, profession TEXT, work_time INTEGER)")
    conn.commit()

    # Читаем данные из файла
    workbook = openpyxl.load_workbook(filename="зп февраль 2023.xlsx")
    worksheet = workbook.active

    data = []
    for row in worksheet.iter_rows(min_row=5, values_only=True):
        tab_number, name, profession, work_time = row[1], row[2], row[3], round(row[6] or 0)
        if tab_number:
            data.append((tab_number, name, profession, work_time))

    # Сохраняем данные в базу данных
    for row in data:
        cursor.execute("INSERT INTO mytable (tab_number, name, profession, work_time) VALUES (?, ?, ?, ?)",
                       row)
    conn.commit()
    conn.close()


def update_work_time():
    """Записываем время работы с базы данных в файл"""
    # подключаемся к базе данных
    conn, cursor = opening_the_database()

    # получаем все записи из таблицы базы данных
    cursor.execute("SELECT tab_number, work_time FROM mytable")
    rows = cursor.fetchall()

    # загружаем файл Excel и выбираем нужный лист
    workbook = load_workbook('табульки.xlsx')
    sheet = workbook['Май 2023']

    # проходимся по всем записям из базы данных
    for row in rows:
        # ищем соответствующую запись в файле Excel
        for i in range(11, sheet.max_row + 1):
            if sheet.cell(row=i, column=2).value == row[0]:
                # обновляем значение в файле Excel
                sheet.cell(row=i, column=7).value = row[1]
                print(f"Updated value for tab_number {row[0]} to {row[1]}")
                break

    # сохраняем изменения в файл Excel
    workbook.save('табульки.xlsx')

    # закрываем соединение с базой данных
    conn.close()


def main():
    we_form_the_working_hours()
    update_work_time()


main()
