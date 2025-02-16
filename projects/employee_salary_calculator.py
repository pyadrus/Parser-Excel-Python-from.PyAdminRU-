import sqlite3
from tkinter import *
from tkinter.filedialog import askopenfilename

from openpyxl import load_workbook


def main():
    print('1 - Парсинг списка сотрудников\n'
          '2 - Подсчет зарплаты\n')
    user_input = int(input('Выберите пункт меню: '))
    if user_input == 1:
        print('Парсинг списка сотрудников')
    elif user_input == 2:
        my_salary()
    elif user_input == 3:
        print('Подсчет зарплаты')
        f_i_o = 'Филищинский С.А.'
        chasov = 108.22
        chas = 159
        klass = 25 * (chasov * chas / 100)
        visluga = (1.09 * (chasov * chas)) - (chasov * chas)
        print(f'ФИО: {f_i_o}, Зарплата: {chasov * chas + klass + visluga}')


def opening_a_file():
    """Окно для выбора файла Excel"""
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    return filename


def opening_the_database():
    """Открытие базы данных"""
    conn = sqlite3.connect('../data.db')  # Создаем соединение с базой данных
    cursor = conn.cursor()
    return conn, cursor


def po_parsing_jul_2023():
    """Изменение от 19.01.2024 Парсинг май 2023"""

    conn, cursor = opening_the_database()
    filename = opening_a_file()  # Открываем выбор файла Excel для чтения данных
    workbook = load_workbook(filename=filename)  # Загружаем выбранный файл Excel
    sheet = workbook.active
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS baza_processing (service_number, fio, profession, district, zp)''')
    # Считываем данные из колонок A и H и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=6, max_row=1100, values_only=True):
        service_number = str(row[5])  # Преобразуем значение в строку
        zp = str(row[10])  # Преобразуем значение в строку
        fio = str(row[6])  # Преобразуем значение в строку
        profession = str(row[3])  # Преобразуем значение в строку
        district = str(row[1])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute('SELECT * FROM baza_processing WHERE service_number = ?', (service_number,))
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute('INSERT INTO baza_processing VALUES (?, ?, ?,?,?)', (service_number, zp,))
        # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


def my_salary():
    """Моя ЗП"""
    working_days_in_january = 17  # Рабочие дни
    working_days_in_february = 20  # Рабочие дни

    """ от 1 года до 2 лет, от 2 лет до 3 лет, от 3 лет до 5 лет, от 5 лет до 10 лет
        от 10 лет до 15 лет, от 15 до 20 лет, от 20 до 25 лет, свыше 25 лет """

    koef_visl = {'Подземных работы': [6.7, 8.3, 11.7, 13.3, 15.0, 16.7, 23.0, 25.0],
                 'В технологической цепочке': [4.2, 5.8, 8.3, 10.0, 11.7, 13.3, 18.3, 19.9],
                 'Строительно-производственный персонал': [3.3, 5.0, 6.7, 8.3, 10.0, 12.5, 17.2, 18.7],
                 'Другие работники предприятия': [2.5, 4.2, 5.8, 7.5, 9.2, 10.8, 14.9, 16.7]}

    # Виды начислений
    oklad = 41074  # Оклад
    chasovaia_tarif_stavka = 0  # Часовая тарифная ставка
    vicluga_let = 26  # Выслуга лет (отработано на предприятии)
    professia = 'Начальник ООТ и ЗП'
    nochnie = 0  # Ночные
    vicher = 0  # Вечерние
    doplata_do_mrot = 0  # Доплата до МРОТ
    pererabotka_chasov = 0  # Переработка часов
    doplata_za_rashirenie_obsl = 0  # Доплата за расширение обслуживания
    otpusknie = 0  # Отпускные
    doplata_za_nenormir_vodit = 0  # Доплата за не нормированный день водителям

    if professia == 'Начальник ООТ и ЗП':
        if vicluga_let >= 1 and vicluga_let <= 2:
            sum_visl_let = (oklad / 100) * koef_visl['В технологической цепочке'][0]
        if vicluga_let >= 2 and vicluga_let <= 3:
            sum_visl_let = (oklad / 100) * koef_visl['В технологической цепочке'][1]
        if vicluga_let >= 3 and vicluga_let <= 5:
            sum_visl_let = (oklad / 100) * koef_visl['В технологической цепочке'][2]
        if vicluga_let >= 5 and vicluga_let <= 10:
            sum_visl_let = (oklad / 100) * koef_visl['В технологической цепочке'][3]
        if vicluga_let >= 10 and vicluga_let <= 15:
            sum_visl_let = (oklad / 100) * koef_visl['В технологической цепочке'][4]
        if vicluga_let >= 15 and vicluga_let <= 20:
            sum_visl_let = (oklad / 100) * koef_visl['В технологической цепочке'][5]
        if vicluga_let >= 20 and vicluga_let <= 25:
            sum_visl_let = (oklad / 100) * koef_visl['В технологической цепочке'][6]
        if vicluga_let >= 25:
            sum_visl_let = (oklad / 100) * koef_visl['В технологической цепочке'][7]

    # Подсчет заработной платы
    zp = (oklad + sum_visl_let)  # Зарплата ФОТ

    print(f'Рабочие дни: {working_days_in_january}')
    print(f'Оклад: {oklad}')
    print(f'Выслуга лет: {sum_visl_let}')
    print(f'Заработная плата работника: {zp}\n')


if __name__ == '__main__':
    main()
