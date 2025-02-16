import re

from openpyxl import load_workbook
from rich import print

workbook = load_workbook(filename='Перечень ОНИ Минстрой (для Даши) Захаров.xlsx')
sheet = workbook.active

current_row = 5  # Start from the 3rd row

for row in sheet.iter_rows(min_row=5, max_row=483, values_only=True):
    number = str(row[14])  # Считываем значение в колонке
    print(number)

    matches = re.findall(r'Инв\.№(\d+)', number)

    if matches:
        result = matches[0]
        print(result)
        # Set the value of the cell in the 16th column, current_row
        sheet.cell(row=current_row, column=16, value=result)
    else:
        print("Не удалось найти номер инвентаря.")

    current_row += 1  # Увеличиваем номер строки

# Сохраняем изменения в файле
workbook.save(filename='Перечень ОНИ Минстрой (для Даши) Захаров.xlsx')
