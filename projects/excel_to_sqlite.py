import sqlite3

import openpyxl


def extract_colored_cells_and_save_to_db(excel_file, sheet_name, db_file, table_name):
    """
    Извлекает значения из закрашенных ячеек столбца B указанного листа Excel-файла
    и сохраняет их в таблицу SQLite базы данных.

    Параметры:
    - excel_file (str): Путь к файлу Excel.
    - sheet_name (str): Название листа в Excel-файле.
    - db_file (str): Путь к файлу базы данных SQLite.
    - table_name (str): Название таблицы в базе данных SQLite.

    Логика работы:
    1. Открывает указанный файл Excel и выбирает нужный лист.
    2. Проходит по всем ячейкам в столбце B и собирает значения из закрашенных ячеек.
    3. Сохраняет эти значения в указанную таблицу SQLite.
    """
    try:
        # Загружаем файл Excel
        workbook = openpyxl.load_workbook(excel_file)

        # Выбираем нужный лист
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден в файле Excel.")
        sheet = workbook[sheet_name]

        # Создаем список для хранения значений из закрашенных ячеек
        values = []

        # Проходим по всем строкам в столбце B
        for row in sheet.iter_rows(min_row=1, min_col=2, max_col=2):
            for cell in row:
                # Проверяем, является ли ячейка закрашенной
                if cell.fill.start_color.index != '00000000':  # '00000000' означает прозрачный цвет
                    if cell.value is not None:  # Проверяем, что значение не пустое
                        try:
                            # Преобразуем значение в целое число
                            value = int(cell.value)
                            values.append(value)
                        except ValueError:
                            print(f"Пропущено значение: {cell.value} (не является числом).")

        # Выводим список извлеченных значений
        print("Извлеченные значения:", values)

        # Подключаемся к базе данных SQLite
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()

        # Создаем таблицу, если её нет
        cursor.execute(f'''CREATE TABLE IF NOT EXISTS {table_name} (value INTEGER)''')

        # Вставляем значения в таблицу
        for value in values:
            cursor.execute(f"INSERT INTO {table_name} (value) VALUES (?)", (value,))

        # Сохраняем изменения и закрываем соединение
        conn.commit()
        conn.close()

        print(f"Данные успешно сохранены в таблицу '{table_name}' базы данных '{db_file}'.")

    except FileNotFoundError as e:
        print(f"Ошибка: Файл не найден. ({e})")
    except Exception as e:
        print(f"Произошла ошибка: {e}")


# Пример использования функции
if __name__ == "__main__":
    # Укажите пути к вашим файлам и параметры
    excel_file = 'РАСЧЕТНАЯ-ВЕДОМОСТЬ-за-4 месяца-2023.xlsx'
    sheet_name = 'Апрель 2023'
    db_file = 'database.db'
    table_name = 'data'

    # Вызов функции для обработки данных
    extract_colored_cells_and_save_to_db(excel_file, sheet_name, db_file, table_name)
