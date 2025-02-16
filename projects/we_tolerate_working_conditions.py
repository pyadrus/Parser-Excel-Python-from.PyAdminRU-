import openpyxl

from database.database import opening_the_database

conn, cursor = opening_the_database()

# Читаем файл Excel
wb = openpyxl.load_workbook('табульки.xlsx')
sheet = wb.active

# Проходим по строкам в файле Excel и ищем соответствующие записи в базе данных
for i, row in enumerate(sheet.iter_rows(min_row=12, values_only=True), start=12):
    tabelnyi_nomer = row[2]
    cursor.execute("SELECT percent FROM mytable WHERE tabelnyi_nomer = ?", (tabelnyi_nomer,))
    result = cursor.fetchone()
    if result:
        percent = result[0]
        # Изменяем значение ячейки в столбце Q для текущей строки
        sheet.cell(row=i, column=38).value = percent
        print(f"Обновлено значение ячейки Q{i} на {percent}")

# Сохраняем изменения в файле Excel
wb.save('табульки.xlsx')
print("Сохранение выполнено успешно!")

# Закрываем соединение с базой данных
conn.close()
print("Соединение с базой данных закрыто.")
