import json
import sqlite3

import openpyxl
from loguru import logger

month_dict = ['01_2024', '02_2024', '03_2024']
file1_path = ['109', '112', '122', '124', '127']  # Названия файлов


def we_write_data_to_the_database():
    """Записываем данные в базу данных"""
    for m in month_dict:
        logger.info(m)
        conn = sqlite3.connect('your_database.db')  # Открываем соединение с базой данных SQLite
        cursor = conn.cursor()

        # Создаем таблицу, если ее еще нет
        cursor.execute(f'''CREATE TABLE IF NOT EXISTS month_{m} (district, name, employee_id, full_name,
                          profession, salary, set_amount, set_percentage, sum_by_type)''')
        conn.commit()

        try:
            for dist in file1_path:
                print(dist)
                wb1 = openpyxl.load_workbook(f'{m}/{dist}.xlsx')
                sheet1 = wb1['1']  # Открываем вкладку

                # Получаем значения из первого файла, начиная с 4 строки, столбец E (исправлено на 5)
                for i in range(3, sheet1.max_row + 1):
                    cell_value1 = str(sheet1.cell(row=i, column=1).value)  # Наименование участка
                    cell_value2 = str(sheet1.cell(row=i, column=2).value)  # Таб.№
                    cell_value3 = str(sheet1.cell(row=i, column=3).value)  # ФИО
                    cell_value6 = str(sheet1.cell(row=i, column=6).value)  # Профессия
                    cell_value8 = str(sheet1.cell(row=i, column=8).value)  # Тариф/оклад
                    cell_value11 = str(sheet1.cell(row=i, column=11).value)  # Установленная сумма
                    cell_value12 = str(sheet1.cell(row=i, column=12).value)  # Установленный %
                    cell_value13 = str(sheet1.cell(row=i, column=13).value)  # Сумма по виду:

                    print(f'Вид оплаты: {dist}, Наименование участка: {cell_value1}, Таб.№: {cell_value2}, '
                          f'ФИО: {cell_value3}, Профессия: {cell_value6}, Тариф/оклад: {cell_value8}, '
                          f'Установленная сумма: {cell_value11}, Установленный %: {cell_value12}, '
                          f'Сумма по виду: {cell_value13}')

                    # Вставляем данные в базу данных
                    cursor.execute(f'''INSERT INTO month_{m} (district, name, employee_id, full_name, profession, 
                                                             salary, set_amount, set_percentage, sum_by_type)
                                                             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
                        dist, cell_value1, cell_value2, cell_value3, cell_value6, cell_value8, cell_value11,
                        cell_value12,
                        cell_value13))

            conn.commit()

            # Закрываем соединение с базой данных
            conn.close()
        except Exception as e:
            logger.error(e)


def read_data_from_the_database():
    """Читаем данные из базы данных"""
    # Открываем соединение с базой данных SQLite
    conn = sqlite3.connect('your_database.db')
    cursor = conn.cursor()  # Создаем курсор
    cursor.execute('SELECT * FROM month_01_2024')
    rows = cursor.fetchall()  # Выбираем все данные из таблицы
    conn.close()
    print(rows)  # Выводим данные из таблицы в консоль.
    return rows  # Возвращаем данные из таблицы


def read_list(rows):
    """Читаем данные из базы данных"""
    dict_list = []  # Создаем пустой список
    # Выводим результаты в консоль
    for row in rows:
        dict_list.append(row[8])  # Добавляем данные в список dict_list
    return dict_list  # Возвращаем данные из таблицы в консоль


def read_data_base(month, district_search):
    """Читаем данные из базы данных"""

    conn = sqlite3.connect('your_database.db')  # Открываем соединение с базой данных SQLite
    cursor = conn.cursor()  # Создаем курсор
    query = f"SELECT * FROM {month} WHERE district = '{district_search}'"  # Формируем SQL-запрос для поиска
    cursor.execute(query)  # Выполняем SQL-запрос
    rows = cursor.fetchall()  # Получаем результаты запроса
    conn.close()  # Закрываем соединение с базой данных
    return rows


def data_analysis():
    """Анализ данных"""

    district_search = ['109', '112', '122', '124', '127']  # Названия файлов
    month = ['month_01_2024', 'month_02_2024', 'month_03_2024']  # Названия файлов

    for dis in district_search:
        results = {}  # Создаем словарь для хранения результатов

        for m in month:
            rows = read_data_base(m, dis)  # Получаем данные из базы данных по данному столбцу 'district'
            dict_list = read_list(rows)  # Выводим результаты в консоль

            for i, number in enumerate(dict_list):
                if not isinstance(number, int):
                    dict_list[i] = float(number)

            total = 0
            for number in dict_list:
                total += number

            print(f'Cумма по виду {dis} за {m}: {round(total, 2)}')

            if m == 'month_01_2024':
                ms = '01.2024'
                results[ms] = round(total, 2)  # Добавляем сумму в словарь
            elif m == 'month_02_2024':
                ms = '02.2024'
                results[ms] = round(total, 2)  # Добавляем сумму в словарь
            elif m == 'month_03_2024':
                ms = '03.2024'
                results[ms] = round(total, 2)  # Добавляем сумму в словарь

        with open(f'results{dis}.json', 'w') as f:
            json.dump(results, f)  # Записываем результаты в файл


def main():
    """Главная функция"""
    print('[1] Парсинг данных\n'
          '[2] Анализ данных\n')
    user_input = input('Выберите пункт меню: ')
    if user_input == '1':
        we_write_data_to_the_database()  # Записываем данные в базу данных SQLit
    elif user_input == '2':
        data_analysis()  # Анализ данных в консоли


if __name__ == '__main__':
    main()  # Запускаем функцию main()
