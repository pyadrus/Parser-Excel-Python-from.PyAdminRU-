from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def main():
    # Открываем выбранный файл Excel для чтения данных
    workbook = load_workbook(filename='Шаблон ОДИ испр. (МУЭ тлг.5463) 30.10.23 (рабоч. черновик).xlsx')
    sheet = workbook.active

    # Создаем стиль для выделения красным цветом
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    for row in sheet.iter_rows(min_row=2, max_row=1224, values_only=True):
        number = str(row[5])  # Преобразуем значение в строку Инвентарный номер

        parts = number.split()

        for part in parts:
            modified_string = part.replace(";", "")
            # print(modified_string)

            if modified_string == "000781388":
                print("Ура, нашли 000781387")
                cost_000781450 = 7425.58
                cost = float(row[3])  # Преобразуем значение в число
                cost_ost = cost - cost_000781450
                print(cost_ost)

                # Обновляем значение в третьей колонке (индекс 2) в текущей строке
                row[2] = cost_ost

                # Применяем стиль к ячейке, чтобы выделить ее красным цветом
                sheet.cell(row=row[0], column=3).fill = red_fill

    # Сохраняем изменения в файле Excel для записи результатов
    workbook.save(filename='Шаблон ОДИ испр. (МУЭ тлг.5463) 30.10.23 (рабоч. черновик).xlsx')
    workbook.close()


main()
