import os

import pyexcel as p
from openpyxl import load_workbook

for x in os.listdir():
    if x.endswith('.xls'):
        file = os.path.splitext(x)[0]
        print(f"[bold green]Найденные файлы: {file}.xls")

        # Конвертируем файл xls в xlsx, так как библиотека openpyxl не работает с xls
        p.save_book_as(file_name=f'{file}.xls',
                       dest_file_name=f'{file}.xlsx')

        try:
            # Находим файл в корневой папке
            wb = load_workbook(filename=f'{file}.xlsx')
            # Открываем 1 рабочую вкладку
            sheet_ranges = wb['1']
            # Считываем строку где написан код участка
            file_new = sheet_ranges['P7'].value
            print(file_new)
            os.rename(f"{file}.xls", f"{file_new}.xls")



        except KeyError:
            # Если нет рабочей вкладки под именем 1, то находим п2 (в простое) 
            wb = load_workbook(filename=f'{file}.xlsx')
            sheet_ranges = wb['п2']
            # Считываем строку где написан код участка
            file_new = sheet_ranges['P7'].value
            print(file_new)
            os.rename(f"{file}.xls", f"{file_new}.xls")

input("Нажмите ENTER")
